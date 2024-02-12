import datetime
from os import name
from matplotlib import dates
from openpyxl import load_workbook
import calendar
import pandas as pd

file_path = 'data4bkping/MonthlyTimesheet.xlsx'

# Load the provided Excel file using openpyxl
wb = load_workbook(file_path)
ws = wb.active



def get_dates_wkdays(year, month):
    """
    Returns the list of dates and corresponding weekdays for the given year and month
    """
    num_days = calendar.monthrange(year, month)[1]
    days = [i for i in range(1, num_days + 1)]
    weekdays = [calendar.day_name[(calendar.weekday(year, month, day))] for day in days]
    return days, weekdays

# Creating a new dataframe for the May 2022 timesheet
# First, we'll create the header for the dates and corresponding weekdays
#num_days = calendar.monthrange(year, month)[1]
#days = [i for i in range(1, num_days + 1)]
#weekdays = [calendar.day_name[(calendar.weekday(year, month, day))] for day in days]



#wkdays_of_mon_year = [calendar.day_name[(calendar.weekday(2022, 5, day))] for day in range(1, 32)]
#dates_of_mon_year = [day for day in range(1, 32)]



def update_dates_and_weekdays(worksheet, start_row, date_list, weekday_list, month, year):
    month_name = calendar.month_name[month]
    worksheet.cell(6,3).value = datetime.datetime(year,month,date_list[-1])
    # clear the cells first
    for i in range(len(date_list)):
        date_cell = worksheet.cell(row=start_row, column=4 + i)
        date_cell.value = ''
    for i in range(len(date_list)):
        # Assuming the dates start from the 4th column (index 3) as observed in the template
        date_cell = worksheet.cell(row=start_row, column=4 + i)
        weekday_cell = worksheet.cell(row=start_row + 1, column=4 + i)

        #weekday_cell.value = weekday_list[i]

        # Update values if not Saturday , Sunday or Monday
        if weekday_list[i] in ['Saturday', 'Sunday']:
            date_cell.value = ''
            continue
        if weekday_list[i] in ['Monday']:
            date_cell.value = 0
            continue
        # Update values for other days
        date_cell.value = 3.00




def gen_timesheet(year, month):
    dates_of_mon_year, wkdays_of_mon_year = get_dates_wkdays(year, month)
    # Updating the worksheet with May 2022 dates and weekdays
    update_dates_and_weekdays(ws, 10, dates_of_mon_year[:15], wkdays_of_mon_year[:15],month,year)
    update_dates_and_weekdays(ws, 18, dates_of_mon_year[15:], wkdays_of_mon_year[15:],month,year)
    month_name = calendar.month_name[month]
    output_file_path_excel_format_updated = f'data4bkping/{year}_{month_name}_Timesheet.xlsx'
    wb.save(output_file_path_excel_format_updated)
    #totla_day = ws.cell(27,18).value
    #print(f"Total day: {totla_day}")

def main():
    for month in range(4, 13):
        #gen_timesheet(2022, month)

        get_total_days(2022,month)

def get_total_days(year, month):
    import xlwings as xw
    month_name = calendar.month_name[month]
    file_path_excel = f'data4bkping/{year}_{month_name}_Timesheet.xlsx'
    #wb = load_workbook(file_path_excel, data_only=True)
    #ws = wb.active
    wbxl=xw.Book(file_path_excel)
    #day_cell = ws.cell(27,18)
    day_cell = wbxl.sheets['Subcontractor Monthly Timesheet'].range('R27')
    totla_day = day_cell.value
    print(f"Total day: {totla_day}")


    #gen_timesheet(2022, 6)

if __name__ == '__main__':
    main()




